# Databricks notebook source
from os import path
import pandas as pd
import csv
import os.path
import numpy as np
import string
from pandas.core.frame import DataFrame
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook

unwanted = string.punctuation
stage_dir = "/dbfs/mnt/sgre-apps/data_validation_engine/"
stage_file = "/dbfs/mnt/sgre-apps/data_validation_engine/error_Log.csv"
#print(unwanted)
# quit()
err_summary = {}
err_summary["ruleId"] = []
err_summary["colName"] = []
err_summary["errorDesc"] = []
err_summary["validation"] = []
err_summary["errorValue"] = []
err_summary["record"] = []

def getDFfromCsv(csvPath, skip_rows):
  try:
    
     flag = os.path.exists(csvPath)
     if flag == True:
        df = pd.read_csv(csvPath, skiprows=int(skip_rows), dtype=object)
        return df
  except:
        raise

def getDFfromXls(xlsPath, skip_rows):
    flag = os.path.exists(xlsPath)
    if flag == True:
        #    wb = load_workbook(filename=xlsPath)
        #    ws = wb['towers']
        #    data = ws.values
        #    columns = next(data)[0:]
        #    df = pd.DataFrame(data, columns=columns)
        df = pd.read_excel(xlsPath, sheet_name=0, header=int(skip_rows), dtype=object)
        return df




def getDFfromXlsx(xlsxPath, sheet_name,Column_address,Column_address1,SKIP_ROWS):
    flag = os.path.exists(xlsxPath)
    if flag == True:
        # df = pd.read_excel(xlsxPath, engine='openpyxl', skiprows=skip_rows, dtype=object)
        # df = df.iloc[2:8,3:6]
        # print("df",df)
        wb = load_workbook(filename=xlsxPath, read_only=True)

        ws = wb[sheet_name]
        data_rows = []
        for row in ws[Column_address:Column_address1]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)

        df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
        #print(df)
    #     isHeaderOn = "true"
    #     isInferSchemaOn = "false"
    #     #sheet address in excelpip install pyspark

    #     sampleAddress = "'towers'!A4"
    #     #read excelfile
    #     #df = spark.read.format("com.crealytics.spark.excel").option("header", isHeaderOn).option("inferSchema", isInferSchemaOn).option("treatEmptyValuesAsNulls", 'True').option("dataAddress", sampleAddress).load(xlsxPath)
    #     spark = SparkSession.builder.appName("ExcelRead").getOrCreate()

    #     df = spark.read.format("com.crealytics.spark.excel")\
    #   .option("useHeader", "true")\
    #   .option("treatEmptyValuesAsNulls", "true")\
    #   .option("inferSchema", "true")\
    #   .option("addColorColumns", "False")\
    #   .load(xlsxPath)
        
        return df


def unmerge_excel_files(fileName, skip_rows):
    # Function to take a merged cell as input and return the unmerged version
    def unmerge_cells_predif(min_col, min_row, max_col, max_row):
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        print("top left coluymn",top_left_cell_value)
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value

    skiprows = skip_rows
    #SRC_DIR = '/dbfs/mnt/sgre-internal/validation_engine/Source/'
    SRC_DIR = 'C:/rulengine_master/Report'
    #stage_dir = "/dbfs/mnt/sgre-apps/data_validation_engine/"
    #stage_dir = 'C:/rulengine_master/'
    
    #OUT_FILE = stage_dir + fileName + ".csv"
    OUT_FILE =  fileName + ".csv"
    

    fileNamee = fileName + ".xlsx"
    xlsx_file = Path(SRC_DIR, fileNamee)
    wb_obj = load_workbook(xlsx_file)
    sheet = wb_obj.active
    dlt = []

    for cell_group in sheet.merged_cells.ranges:
        i = 0
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        dlt.append(range_boundaries(str(cell_group)))

    for i in dlt:
        if i[3] <= skiprows:
            unmerge_cells_predif(i[0], i[1], i[2], i[3])

    index_row = []

    for i in range(1, sheet.max_row):
        if sheet.cell(i, 1).value is None:
            index_row.append(i)
  
    df = pd.DataFrame(OUT_FILE)
   
    #df.to_csv('/dbfs/mnt/sgre-internal/validation_engine/Source/merge_unmerge.csv', index=False, header=False)
    df.to_csv(OUT_FILE, index=True, header=True)
    print("OUT file",OUT_FILE)
    print("mmydf",df)

    return df

def getDFfromXlsxMerge(xlsxPath, skip_rows):
    flag = os.path.exists(xlsxPath)
    if flag == True:
        fileName = xlsxPath.split('/')[-1].split('.')[0]
        print(xlsxPath)
        print("file name---->>>>>>",fileName)
        skip_rows=1
        df= unmerge_excel_files(fileName, int(skip_rows))
        #print("csvpath---->>>",csvPath)
        
        #df = getDFfromCsv(csvPath, skip_rows)
        
        return df



def GetColumnDataType(datafram, colName):
    if colName in datafram.columns:
        datatypes = datafram.dtypes[colName]
        return datatypes





def check_dtype(DataFram, colName):
    try:
    #    if pd.to_datetime(DataFram[colName].str.split(' ').str[0], format='%y-%m-%d').notnull().all()  :
    #        return "date"
    # except:
    #    pass
        if DataFram[colName].eq('').all() or pd.isnull(DataFram[colName]).all():
            print("pass")
            return "pass"
        elif DataFram[colName].dropna().isin(['Yes', 'No']).all():
            return "bool"
        elif DataFram[colName].dropna().astype(str).str.contains(pat='[a-zA-Z/-]', regex=True).any():
            return "string"
        # elif DataFram[colName].dropna().astype(str).str.contains(pat='[a-zA-Z0-9/-]', regex=True).any():
        #     return "string"
        # if DataFram[colName].dropna().astype(str).str.contains(r'[@#&$%+/^*]').any():
        #    print("Is Not String")
        # else:
        #    print("String")
        # if (DataFram[colName].fillna(-9999) % 1 ==0).all():
        #    return "int"
        #try:

        elif (DataFram[colName].str.replace('.', '', regex=True).fillna(-9999).astype(int)).all():
                if (DataFram[colName].dropna().astype(float) % 1 != 0.0).any():
                    return "float"
                else:
                    return "int"
    except:
        pass


# Method to perform business validation
# Method to perform business validation
def check_ruleValidation(DataFram, colName, Rule_id, validation_operator, value_to_be_match):

    stage_file = "C:/rulengine_master/Report/error/error_Log.csv"
    var1 = "Pass"

    if 'Not Null' in validation_operator:        

        for jdict in DataFram[DataFram[colName].notnull() == False].to_dict(orient='records'):

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Not Null Constraint Failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append('')
            err_summary["record"].append(jdict)
            var1 = "Fail"
           
    elif 'Null' in validation_operator:

        for jdict in DataFram[DataFram[colName].isnull() == False].to_dict(orient='records'):

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Null Constraint Failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append('')
            err_summary["record"].append(jdict)
            var1 = "Fail"

    elif "Regular expression" in validation_operator:    

         for jdict in DataFram[
            DataFram[colName].astype(str).str.contains(value_to_be_match).astype(str) == 'False'].to_dict(orient='records'):

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Regex validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"


    elif 'Contains' in validation_operator:            

        for jdict in DataFram[
            DataFram[colName].astype(str).str.contains(value_to_be_match).astype(str) == 'True'].to_dict(orient='records'):

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("contains validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"
   
    elif "Greater than" in validation_operator:  

        for jdict in DataFram[DataFram[colName].astype(int) > int(value_to_be_match)].to_dict(orient='records'):        

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Greater than validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"

    elif "Less than" in validation_operator:   

      for jdict in DataFram[DataFram[colName].astype(int) < int(value_to_be_match)].to_dict(orient='records'):   

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Less than validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"        

    elif "Equals to" in validation_operator:     

      for jdict in DataFram[DataFram[colName] == value_to_be_match].to_dict(orient='records'):    

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Equals To validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"    
    
    elif "Does not equals to" in validation_operator:       

      for jdict in DataFram[DataFram[colName] != value_to_be_match].to_dict(orient='records'):       

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Does Not Equals to validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"
    
    elif "Year" in validation_operator:     

      for jdict in DataFram[(DataFram[colName].astype(int).between(1900,2200,inclusive=True)).astype(str) == 'False'].to_dict(orient='records'): 

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("year validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"

    elif "Boolean" in validation_operator:       

      for jdict in DataFram[(DataFram[colName].astype(str).str.upper().str.strip().isin(['TRUE','FALSE','0','1'])).astype(str) == 'False'].to_dict(orient='records'): 
            print(jdict)
            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Boolean validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"
    
    # elif "date" in validation_operator:     

    #   for jdict in DataFram[(DataFram[colName].astype(int).between(01,31,inclusive=True)).astype(str) == 'False'].to_dict(orient='records'): 

    #         err_summary["ruleId"].append(Rule_id)
    #         err_summary["colName"].append(colName)
    #         err_summary["errorDesc"].append("year validation failed")
    #         err_summary["validation"].append(value_to_be_match)
    #         err_summary["errorValue"].append(jdict[colName])
    #         err_summary["record"].append(jdict)
    #         var1 = "Fail"

    
    elif 'Date' in validation_operator or 'DateTime' in validation_operator :

      for jdict in DataFram[
            DataFram[colName].astype(str).str.contains(value_to_be_match).astype(str) == 'False'].to_dict(orient='records'):

            err_summary["ruleId"].append(Rule_id)
            err_summary["colName"].append(colName)
            err_summary["errorDesc"].append("Date validation failed")
            err_summary["validation"].append(value_to_be_match)
            err_summary["errorValue"].append(jdict[colName])
            err_summary["record"].append(jdict)
            var1 = "Fail"

    elif 'None' in validation_operator:

        # for jdict in DataFram[DataFram[colName].astype(str).str.contains(value_to_be_match).astype(str) == 'True'].to_dict(orient='records'):       
             pass
            
            # err_summary["ruleId"].append('')
            # err_summary["colName"].append('')
            # err_summary["errorDesc"].append('')
            # err_summary["validation"].append('')
            # err_summary["errorValue"].append('')
            # err_summary["record"].append('')
    
    else:
        var1="pass"

    err_out = pd.DataFrame(err_summary)
    # print(err_out)
    err_out.to_csv(stage_file, index=False)
    return var1


# COMMAND ----------

# xlsxPath='/dbfs/mnt/sgre-internal/validation_engine/Source/MASTER_TOWER_COMPONENTS.xlsx'
# skip_rows=4
# df=getDFfromXlsxMerge(xlsxPath, skip_rows)
# print(df)
#df = getDFfromCsv('C:/rulengine_master/sample3.csv',0)
#print(check_ruleValidation(df,"month","1","Greater than","2"))
# print(df['maximum_section_weight_kg'])
# print(check_dtype(df, 'maximum_section_weight_kg'))
# print(df['date_of_hit'].str.split(' ').str[0])
# print(pd.to_datetime(df['date_of_hit'].str.split(' ').str[0], infer_datetime_format=True))
# print(pd.to_datetime(df['date_of_hit'].str.split(' ').str[0], format='%y-%m-%d'))
# absDataFilePath='/dbfs/mnt/sgre-internal/validation_engine/Source/towersxlhdr.xlsx'
# sampleDataFilePath='/dbfs/mnt/sgre-internal/validation_engine/Source/towersxlhdr.xlsx'
# df = getDFfromXlsx(sampleDataFilePath, 3)
#print(check_dtype(df, 'Variable_category'))
# print(df['noise_mode_n1_old'])
# print(df['hub_height_m'])
# df1 = df.fillna("",inplace=True)
# print(check_dtype(df, 'hub_height_m'))
