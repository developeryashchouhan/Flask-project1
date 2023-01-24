from openpyxl import load_workbook

wb = load_workbook(filename='Vdata.xlsx', 
                   read_only=True)

ws = wb['sheet1']

# Read the cell values into a list of lists
data_rows = []
for row in ws['A4':'D10']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe
import pandas as pd
#df = pd.DataFrame(data_rows,header=True)

df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
print(df)